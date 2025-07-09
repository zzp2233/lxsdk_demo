# -*- coding: utf-8 -*-
import sys
import re
import datetime
from pathlib import Path
from openpyxl import load_workbook
from pypinyin import pinyin, Style
import libs.MakeFontBinFile as FontBin
import libs.TableDeformation as TblDeform

saved_stdout = sys.stdout
saved_stderr = sys.stderr
global lastErrInfo
global log_file

langList = []

column_list = []
column_list.extend([chr(ord('A') + i) for i in range(26)])
column_list.extend(["A" + chr(ord('A') + j) for j in range(26)])
column_list.extend(["B" + chr(ord('A') + j) for j in range(26)])

def WriteLog(log_str):
    global lastErrInfo
    lastErrInfo = log_str


def GenarateRemark():
    now = datetime.datetime.now()
    remark = """/**
*@file
*@brief utf8 words list
*@details
*@author
*@date   %s
*@version  UTE WordListTool,Version 2.2.1.0
*/

#ifndef _UTE_MULTI_LANGUAGE_H_
#define _UTE_MULTI_LANGUAGE_H_
\n#include <include.h>\n""" % (now.strftime("%Y-%m-%d %H:%M:%S"))
    print(remark)


def RedirectPrint(out_full_path):
    global log_file
    WriteLog("Begin RedirectPrint.")

    # 打开文件，以追加模式写入
    log_file = open(out_full_path, "w")

    # 将标准输出流和错误输出流重定向到文件中
    sys.stdout = log_file
    sys.stderr = log_file

    # 检查当前 sys.stdout 的编码
    if sys.stdout.encoding.lower() != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')


def CloseRedirectPrint():
    global log_file
    # 恢复标准输出流和错误输出流
    sys.stdout = saved_stdout
    sys.stderr = saved_stderr

    # 关闭文件对象
    log_file.close()


def CheckExcelFormat(xlsFullFilePath):
    # 设置语言环境
    # 将 PYTHONIOENCODING 环境变量设置为 utf-8
    # os.environ['PYTHONIOENCODING'] = 'utf-8'
    if sys.stdout != None:
        sys.stdout.reconfigure(encoding='utf-8')

    try:
        # 读取 Excel 文件
        df = load_workbook(xlsFullFilePath)
        sht = df.worksheets[0]

        if (sht['A2'].value == 0 and
                sht['C2'].value == r'中文_2' and
                'AR_5' in sht['F2'].value and
                r'KO_47' in sht['AV2'].value):
            print("验证通过")
            return True

        print("验证失败")
        return False

    except TypeError as e:
        print("[error] open input.xlsx err：", str(e))
        return False



def GenerateWordsMarco(xlsFullFilePath, inputList):
    # 设置语言环境
    # 将 PYTHONIOENCODING 环境变量设置为 utf-8
    # os.environ['PYTHONIOENCODING'] = 'utf-8'
    if sys.stdout != None:
        sys.stdout.reconfigure(encoding='utf-8')

    # 读取 Excel 文件
    dfModule = load_workbook(xlsFullFilePath)
    sht = dfModule.worksheets[0]

    # 将前三列数据保存为列表
    hasId = False
    if sht['D2'].value.count(r'词条ID') > 0:
        hasId = True

    column_ModuleWithWords = sht.iter_rows(min_col=1, max_col=4, min_row=3, values_only=True)
    dupIdx = 0
    sModule = ''
    lstRes = []
    word_number = 0

    langList.clear()

    print('enum \n{')

    for index, it in enumerate(column_ModuleWithWords):
        if it[1] != '' and it[1] is not None:
            sModule = it[1].replace("\r", "").replace("\n", "")
        if it[2] == '' or it[2] is None:
            sWords = ''.replace("\r", "").replace("\n", "")
            if it[0] == '' or it[0] is None:
                word_number = index
                break
        else:
            sWords = it[2].replace("\r", "").replace("\n", "")

        # print(sLanguage + " " + sModule + " " + sWords)
        pyModule = pinyin(sModule, style=Style.FIRST_LETTER)  # 设置拼音风格
        pyWords = pinyin(sWords, style=Style.FIRST_LETTER)  # 设置拼音风格

        # 列表转换为字符串
        sTmp = [''.join(map(str, sublist)) for sublist in pyModule]
        spyModule = ''.join(sTmp).upper()

        sTmp = [''.join(map(str, sublist)) for sublist in pyWords]
        spyWords = ''.join(sTmp).upper()

        # 将字符串中非字母的字符替换为下划线
        spyModule = re.sub(r'[^a-zA-Z0-9]+', '_', str(spyModule))
        spyWords = re.sub(r'[^a-zA-Z0-9]+', '_', str(spyWords))

        sRes = spyModule + "_" + spyWords[:16]

        # 删除连续的下划线
        sRes = sRes.replace("__", "_")

        if sRes in lstRes:
            lstRes.append(sRes + str(dupIdx))
            sRes += str(dupIdx)
            dupIdx += 10
        else:
            lstRes.append(sRes)

        # 删除末尾的下划线
        if sRes[-1] == '_':
            sRes = sRes[:-1]

        word_number = index + 1

        words_id = ''
        if it[3] != '' and it[3] != None:
            words_id = re.sub(r'[^a-zA-Z0-9]+', '_', str(it[3])).upper()
            words_id = words_id.replace("__", "_")
            if words_id[-1] == '_':
                words_id = words_id[:-1]

        try:
            if hasId and words_id != '':
                if words_id.count(r'STR_') > 0:
                    print('    %s, // %s' % (words_id, sWords))
                    langList.append(words_id)
                else:
                    print('    STR_%s, // %s' % (words_id, sWords))
                    langList.append('STR_' + words_id)
            else:
                print('    STR_%s, // %s' % (sRes, sWords))
                langList.append('STR_' + sRes)
        except Exception as e:
            print('#error \"%s 表格错误，请检查词条 %s\"' % (e, sWords))

    # print('\n#define UTE_UTE_MULTI_LANGUAGE_TABLE_MAX_ITEM %s\n\n#endif // _UTE_MULTI_LANGUAGE_H_' % (
    #         int(word_number) + 1))

    print('};\n\nextern const char * const *i18n;')
    for lang in inputList:
        lang_array = lang["name"].split()[0].strip()
        print('extern const char * const i18n_%s[];' % (lang_array.lower())) #extern const char * const i18n_zh[];
    print('\n#endif // _UTE_MULTI_LANGUAGE_H_')


def GenerateWordsPublicFile(inputList):
    # 设置语言环境
    if sys.stdout != None:
        sys.stdout.reconfigure(encoding='utf-8')

    print('#include "ute_multi_language.h"\n')

    for lang in inputList:
        lang_array = lang["name"].split()[0].strip()
        print('#include "ute_multi_language_%s.h"\n' % (lang_array.lower())) #extern const char * const i18n_zh[];

def GenerateWordsUtf8Array(xlsFullFilePath, sLanguage, languageDesc, column, rowStart):
    # 设置语言环境
    # 将 PYTHONIOENCODING 环境变量设置为 utf-8
    # os.environ['PYTHONIOENCODING'] = 'utf-8'
    if sys.stdout != None:
        sys.stdout.reconfigure(encoding='utf-8')

    # 读取 Excel 文件
    df = load_workbook(xlsFullFilePath)
    sht = df.worksheets[0]

    column_words = sht[column_list[column]]
    id_column = sht['A']

    # 输出词条utf8 c数组
    now = datetime.datetime.now()
    remark = """/**
*@file
*@brief utf8 words list h file
*@details language: %s
*@author
*@date   %s
*@version  UTE WordListTool,Version 2.2.1.0
*/
""" % (languageDesc, now.strftime("%Y-%m-%d %H:%M:%S"))
    print(remark)

    sTable = '    '
    sNewLine = '\n'
    sBL = '{'
    sBR = '}'
    # strSavePsRam = "const " if savePsRam else ""
    sLanguageLower = sLanguage.lower()
    sLanguageUpper = sLanguage.upper()
    print(
        f'#ifndef _UTE_MULTI_LANGUAGE_{sLanguageUpper}_H_{sNewLine}#define _UTE_MULTI_LANGUAGE_{sLanguageUpper}_H_{sNewLine}{sNewLine}#include "ute_multi_language.h"{sNewLine}{sNewLine}const char * const i18n_{sLanguageLower}[] ={sNewLine}{sBL}')#{sTable}({strSavePsRam}uint8_t[]){sBL}0x00{sBR}, // 0 保留')
    for index, it in enumerate(column_words[rowStart:]):
        sTmp = str(it.value)
        if it.value == "" or it.value is None:
            if index == len(column_words[rowStart:]) - 1:
                break
            elif id_column[index + rowStart].value is None:
                break
            sTmp = " "

        if sTmp.endswith("&#30;"):
            sTmp = sTmp[:-5]
            sRS = ', \x1e\x1e\x1e'
        else:
            sRS = ''

        # sTmpOringin = sTmp  # 暂存原始字符串
        # if languageDesc == "高棉语":
        #     if not khmer3rd:
        #         utf16_hex = TblDeform.km_deformation(bytearray(sTmp.encode('utf-16-be')))
        #         sTmp = utf16_hex.decode('utf-16-be')
        # elif languageDesc == "印地语(印度）":
        #     utf16_hex = TblDeform.hi_deformation(bytearray(sTmp.encode('utf-16-be')))
        #     sTmp = utf16_hex.decode('utf-16-be')

        utf8_hex = ''.join([char.encode('utf-8').hex() for char in sTmp])
        hVal = '\\x'.join(utf8_hex[i:i + 2] for i in range(0, len(utf8_hex), 2))

        if len(hVal) <= 0:
            hVal = '20'

        res = "[" + langList[index] + '] = "\\x' + hVal + sRS + '\\x00"' + ", // [" + str(index + 1) + "]"
        
        # if languageDesc == "高棉语" or languageDesc == "印地语(印度）":
        #     if khmer3rd and languageDesc == "高棉语":
        #         pass
        #     else:
        #         sTmp = sTmpOringin  # 恢复原始字符串
        sTmp = sTmp.replace("\x0d", "\x20").replace("\x0a", "\x20")
        print(f"{sTable} {res} {sTmp}")
    print('};\n\n#endif\n')


def GenerateWordsUtf8ListData(xlsFullFilePath, column, rowStart, languageID):
    # 读取 Excel 文件
    df = load_workbook(xlsFullFilePath)
    sht = df.worksheets[0]

    column_words = sht[column_list[column]]
    id_column = sht['A']

    data = []
    for index, it in enumerate(column_words[rowStart:]):
        sTmp = str(it.value)
        if it.value == "" or it.value is None:
            if index == len(column_words[rowStart:]) - 1:
                break
            elif id_column[index + rowStart].value is None:
                break
            sTmp = " "

        if languageID == "km":
            utf16_hex = TblDeform.km_deformation(bytearray(sTmp.encode('utf-16-be')))
            sTmp = utf16_hex.decode('utf-16-be')
        elif languageID == "hi_rin":
            utf16_hex = TblDeform.hi_deformation(bytearray(sTmp.encode('utf-16-be')))
            sTmp = utf16_hex.decode('utf-16-be')

        utf8_hex = ' '.join([char.encode('utf-8').hex() for char in sTmp])
        separator = ' '
        xxx = str(utf8_hex).replace("\x20", "")
        res = separator.join(xxx[i:i + 2] for i in range(0, len(xxx), 2))
        if len(res) <= 0:
            res = '20'

        # 将十六进制字符串按空格分割成列表
        hex_list = res.split()

        # 将每个十六进制数值转换为整型并存入列表
        integer_list = [int(hex_value, 16) for hex_value in hex_list]
        integer_list.append(0x00)
        data.append(integer_list)
    return data


def GenerateHeaderFile(sLanguage, savePsRam):
    strSavePsRam = "const " if savePsRam else ""
    info = """#ifndef _UTE_MULTI_LANGUAGE_%s_H_
#define _UTE_MULTI_LANGUAGE_%s_H_
#include "ute_multi_language.h"

extern const uint8_t* %ss_%s_table[UTE_UTE_MULTI_LANGUAGE_TABLE_MAX_ITEM];

#endif // _UTE_MULTI_LANGUAGE_%s_H_
""" % (sLanguage.upper(), sLanguage.upper(), strSavePsRam, sLanguage.lower(), sLanguage.upper())
    print(info)


'''
lang_map_col = [
("zh_rCN", 2,  "中文_简体"),
("en_rUS", 4,  "英语_美国"),
("ar_rIL", 5,  "阿拉伯语_以色列"),
("ru_rRU", 6,  "俄语_俄罗斯"),
("tr_rTR", 7,  "土耳其语_土耳其"),
("ms_rMY", 8,  "马来语_马来西亚"),
("ja_rJP", 9,  "日语_日本"),
("bg_rBG", 11, "保加利亚语_保加利亚"),
("cs_rCZ", 12, "捷克语_捷克共和国"),
("de_rDE", 14, "德语_德国"),
("el",     15, "希腊语"),
("es_rES", 16, "西班牙_西班牙"),
("es_rMX", 17, "拉美西_墨西哥"),
("fr_rFR", 19, "法语_法国"),
("hu_rHU", 21, "匈牙利语_匈牙利"),
("in_rID", 22, "印度尼西亚语_印度尼西亚"),
("it_rIT", 23, "意大利语_意大利"),
("iw_rIL", 24, "希伯来语_以色列"),
("nl_rNL", 26, "荷兰语_荷兰"),
("pl_rPL", 27, "波兰语_波兰"),
("pt_rPT", 29, "葡萄牙语_葡萄牙"),
("ro_rRO", 30, "罗马尼亚语_罗马尼亚"),
("sk_rSK", 31, "斯洛伐克语_斯洛伐克"),
("th_rTH", 34, "泰语_泰国"),
("vi_rVN", 36, "越南语_越南"),
("hi_rIN", 38, "印地语_印度"),
("tl_rPH", 43, "塔加路语_菲律宾"),
("fa_rIR", 44, "波斯语_伊朗"),
("bn",     45, "孟加拉语"),
("ko_rKR", 47, "朝鲜语_韩国"),
]
'''


def MainProc(dir_path, xid, languageID, languageDes, langList):
    print(f"Begin MainProc. xid:{xid}, languageID:{languageID}, languageDes:{languageDes}")
    WriteLog(f"MainProc...{languageID}")

    # 输出目录
    base_dir = Path(dir_path).absolute().parent
    output_dir = Path.joinpath(base_dir, "output")
    if not Path(output_dir).exists():
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        # os.makedirs(output_dir)

    # 检验基本数据
    languageDesc = languageDes
    column = xid
    if not Path(dir_path).exists():
        print(dir_path + "本地文件不存在，请确认！")
        sys.exit()

    filePath = dir_path
    tmpFileName = Path.joinpath(output_dir, "ute_multi_language.h")
    if not Path.exists(tmpFileName):
        print("Begin GenarateRemark.")

        # 将print内容输出到文件
        RedirectPrint(tmpFileName)

        # 注释头生成
        GenarateRemark()
        filePath = dir_path

        # 生成宏定义
        GenerateWordsMarco(filePath, langList)
        CloseRedirectPrint()
        print("End GenarateRemark.")

        tmpFileName = Path.joinpath(output_dir, "ute_multi_language.c")
        # 将print内容输出到文件
        RedirectPrint(tmpFileName)
        GenerateWordsPublicFile(langList)
        CloseRedirectPrint()
        print("End GenaratePublicFile.")

    # if genBinFormat:
    #     print("Begin GenarateBinFormat.")
    #     # 由于Flash对文件名长度有严格限制，这里尽量短
    #     fileName = f"{languageID.lower()}.dct"
    #     tmpFileName = Path.joinpath(output_dir, fileName)
    #     WriteLog(f"Gen {tmpFileName}...")
    #     data = GenerateWordsUtf8ListData(filePath, column, 2, languageID.lower())
    #     WriteLog(f"Gen {tmpFileName} data")
    #     FontBin.WriteListToFontBinFile(tmpFileName, data)
    #     WriteLog(f"Gen {tmpFileName} end.")
    #     print("End GenarateBinFormat.")

    # 生成Utf8数组(第二列，第1行开始处理)
    fileName = f"ute_multi_language_{languageID.lower()}.h"
    tmpFileName = Path.joinpath(output_dir, fileName)
    print(f"Begin proc {tmpFileName}.")
    RedirectPrint(tmpFileName)
    GenerateWordsUtf8Array(filePath, languageID, languageDesc, column, 2)
    CloseRedirectPrint()
    print(f"End proc {tmpFileName}.")

    # fileName = "ute_multi_language_%s.h" % (languageID.lower())
    # tmpFileName = Path.joinpath(output_dir, fileName)
    # print(f"Begin proc {tmpFileName}.")
    # RedirectPrint(tmpFileName)
    # GenerateHeaderFile(languageID, savePsRam)
    # CloseRedirectPrint()
    # print(f"End proc {tmpFileName}.")


def get_executable_dir():
    if getattr(sys, 'frozen', False):
        # The script is bundled, so we return the executable's directory path
        return Path(sys.executable).absolute().parent
    else:
        # The script is running in a normal Python environment
        # We return the Python script's directory path
        return Path(Path(__file__).absolute().parent).absolute().parent
